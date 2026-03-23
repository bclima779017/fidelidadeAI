"""Geração de relatório Excel (.xlsx) com formatação condicional e múltiplas abas."""

import os
from datetime import datetime
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")


def generate_report(results: list[dict], rag_metadata: dict | None = None, score_ponderado: float | None = None, suggestions_data: list[dict] | None = None) -> str:
    """Gera o relatório .xlsx com formatação condicional e retorna o caminho do arquivo.

    Args:
        results: Lista de dicts com os resultados da auditoria.
        rag_metadata: Metadados RAG opcionais {total_pages, total_chunks, chunks_per_page}.
        score_ponderado: Score final ponderado já calculado via scoring.calcular_score_ponderado().
        suggestions_data: Lista de sugestões rankeadas do Protocolo First-Claim.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    filepath = os.path.join(OUTPUT_DIR, f"auditoria_{timestamp}.xlsx")

    # Detecta se há coluna de fontes
    has_sources = any("Fontes Consultadas" in r for r in results)

    columns = ["Pergunta", "Resposta Oficial", "Resposta IA", "Score", "Justificativa"]
    if has_sources:
        columns.append("Fontes Consultadas")

    df = pd.DataFrame(results)[columns]
    df["Score"] = pd.to_numeric(df["Score"], errors="coerce").fillna(-1)

    avg_score = df["Score"].mean()
    min_score = df["Score"].min()
    max_score = df["Score"].max()
    min_idx = df["Score"].idxmin() + 1
    max_idx = df["Score"].idxmax() + 1

    num_cols = len(columns)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Resultados", index=False, startrow=1)
        ws = writer.sheets["Resultados"]

        # Título
        end_col = get_column_letter(num_cols)
        ws.merge_cells(f"A1:{end_col}1")
        title_cell = ws["A1"]
        title_cell.value = f"Auditoria GEO — Kípiai — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        title_cell.font = Font(bold=True, size=14, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Estilo do header
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Formatação condicional por score
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100", bold=True)
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        yellow_font = Font(color="9C6500", bold=True)
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006", bold=True)

        for row in range(3, 3 + len(df)):
            score_cell = ws.cell(row=row, column=4)
            score_val = score_cell.value
            if score_val is not None and score_val != "":
                try:
                    score_num = float(score_val)
                except (ValueError, TypeError):
                    continue
                if score_num >= 70:
                    score_cell.fill = green_fill
                    score_cell.font = green_font
                elif score_num >= 50:
                    score_cell.fill = yellow_fill
                    score_cell.font = yellow_font
                else:
                    score_cell.fill = red_fill
                    score_cell.font = red_font
                score_cell.alignment = Alignment(horizontal="center")

        # Larguras de coluna
        col_widths = {"A": 40, "B": 45, "C": 45, "D": 10, "E": 50}
        if has_sources:
            col_widths["F"] = 50
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Bordas finas
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2, max_row=2 + len(df), min_col=1, max_col=num_cols):
            for cell in row:
                cell.border = thin_border
                if cell.column != 4:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Linha de resumo
        summary_row = 3 + len(df) + 1
        ws.merge_cells(f"A{summary_row}:C{summary_row}")
        ws.cell(row=summary_row, column=1).value = "RESUMO DA AUDITORIA"
        ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12, color="1F4E79")

        score_final_label = f"{score_ponderado:.1f}" if score_ponderado is not None else f"{avg_score:.1f}"
        metrics = [
            (summary_row + 1, "Score Final Ponderado", score_final_label),
            (summary_row + 2, "Score Mínimo", f"{min_score} (Pergunta {min_idx})"),
            (summary_row + 3, "Score Máximo", f"{max_score} (Pergunta {max_idx})"),
            (summary_row + 4, "Total de Perguntas", str(len(df))),
            (summary_row + 5, "Erros (score = -1)", str(len(df[df["Score"] == -1]))),
        ]

        if rag_metadata:
            metrics.append((summary_row + 6, "Modo", "RAG Multi-Página"))
            metrics.append((summary_row + 7, "Páginas Indexadas", str(rag_metadata.get("total_pages", 0))))
            metrics.append((summary_row + 8, "Chunks Indexados", str(rag_metadata.get("total_chunks", 0))))

        for row_num, label, value in metrics:
            ws.cell(row=row_num, column=1).value = label
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            ws.cell(row=row_num, column=2).value = value

        # --- Aba de Metadados RAG ---
        if rag_metadata and rag_metadata.get("chunks_per_page"):
            rag_data = []
            for page_url, chunk_count in rag_metadata["chunks_per_page"].items():
                rag_data.append({"URL da Página": page_url, "Chunks": chunk_count})

            df_rag = pd.DataFrame(rag_data)
            df_rag.to_excel(writer, sheet_name="Metadados RAG", index=False, startrow=1)
            ws_rag = writer.sheets["Metadados RAG"]

            # Título
            ws_rag.merge_cells("A1:B1")
            title_cell = ws_rag["A1"]
            title_cell.value = "Metadados RAG — Páginas Crawleadas"
            title_cell.font = Font(bold=True, size=14, color="1F4E79")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Header
            for col in range(1, 3):
                cell = ws_rag.cell(row=2, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            ws_rag.column_dimensions["A"].width = 70
            ws_rag.column_dimensions["B"].width = 15

            # Resumo no final
            summary_rag_row = 3 + len(df_rag) + 1
            ws_rag.cell(row=summary_rag_row, column=1).value = "Total de Páginas"
            ws_rag.cell(row=summary_rag_row, column=1).font = Font(bold=True)
            ws_rag.cell(row=summary_rag_row, column=2).value = rag_metadata.get("total_pages", 0)

            ws_rag.cell(row=summary_rag_row + 1, column=1).value = "Total de Chunks"
            ws_rag.cell(row=summary_rag_row + 1, column=1).font = Font(bold=True)
            ws_rag.cell(row=summary_rag_row + 1, column=2).value = rag_metadata.get("total_chunks", 0)

        # --- Aba de Sugestões First-Claim ---
        if suggestions_data:
            sug_rows = []
            for i, sug in enumerate(suggestions_data, 1):
                sug_rows.append({
                    "#": i,
                    "Iniciativa": sug.get("titulo", ""),
                    "Eixo": sug.get("eixo", ""),
                    "Impacto": sug.get("impacto", "").capitalize(),
                    "Por que": sug.get("por_que", ""),
                    "O que fazer": sug.get("o_que_fazer", ""),
                    "Perguntas afetadas": " · ".join(sug.get("perguntas_afetadas", [])),
                })

            if sug_rows:
                num_sug_cols = 7
                df_sug = pd.DataFrame(sug_rows)
                df_sug.to_excel(writer, sheet_name="Sugestões", index=False, startrow=1)
                ws_sug = writer.sheets["Sugestões"]

                # Título
                end_sug_col = get_column_letter(num_sug_cols)
                ws_sug.merge_cells(f"A1:{end_sug_col}1")
                title_cell = ws_sug["A1"]
                title_cell.value = "Sugestões de Melhoria — Protocolo First Claim"
                title_cell.font = Font(bold=True, size=14, color="1F4E79")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")

                # Header
                for col in range(1, num_sug_cols + 1):
                    cell = ws_sug.cell(row=2, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)

                # Larguras
                sug_widths = {"A": 5, "B": 30, "C": 25, "D": 12, "E": 50, "F": 50, "G": 40}
                for col_letter, width in sug_widths.items():
                    ws_sug.column_dimensions[col_letter].width = width

                # Formatação condicional por impacto (coluna D)
                for row in range(3, 3 + len(df_sug)):
                    imp_cell = ws_sug.cell(row=row, column=4)
                    imp_val = (imp_cell.value or "").lower()
                    if imp_val == "alto":
                        imp_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        imp_cell.font = Font(color="9C0006", bold=True)
                    elif imp_val == "medio":
                        imp_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        imp_cell.font = Font(color="9C6500", bold=True)
                    else:
                        imp_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        imp_cell.font = Font(color="006100", bold=True)
                    imp_cell.alignment = Alignment(horizontal="center")

                # Wrap text
                for row in ws_sug.iter_rows(min_row=2, max_row=2 + len(df_sug), min_col=1, max_col=num_sug_cols):
                    for cell in row:
                        if cell.column not in (1, 4):
                            cell.alignment = Alignment(wrap_text=True, vertical="top")

    return filepath
